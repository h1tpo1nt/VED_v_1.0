import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm import tqdm
import shutil

# ======================================
# Настройки путей и параметров
# ======================================
input_file = "/content/VED/Input/31 группа 2023-2024.xlsx" # Исходный файл
target_folder = "/content/VED/Target_files"                # Шаблоны файлов
output_folder = "/content/VED/Output"                      # Результаты

# Соответствие кодов ТН ВЭД → файл
code_mapping = {
    "2834210000": "01_Калиевая селитра.xlsx",
    "2915120000": "02_Формиаты.xlsx",
    "3102600000": "03_Нитрат кальция.xlsx",
    "3102900000": "03_Нитрат кальция.xlsx",
    "3105902000": "03_Нитрат кальция.xlsx",
    "2834298000": "03_Нитрат кальция.xlsx",
    "3102500000": "05_Нитрит натрия.xlsx",
    "2834100000": "05_Нитрит натрия.xlsx",
    "3105400000": "06_МАФ.xlsx",
    "3105100000": "07_NPK(S) ВРУ.xlsx",
    "3105200000": "07_NPK(S) ВРУ.xlsx",
    "3105201000": "07_NPK(S) ВРУ.xlsx",
    "3105209000": "07_NPK(S) ВРУ.xlsx",
    "3105510000": "07_NPK(S) ВРУ.xlsx",
    "3105590000": "07_NPK(S) ВРУ.xlsx",
    "3105600000": "07_NPK(S) ВРУ.xlsx",
    "3105908000": "07_NPK(S) ВРУ.xlsx",
    "3105909100": "07_NPK(S) ВРУ.xlsx",
    "3105909900": "07_NPK(S) ВРУ.xlsx",
    "2833210000": "08_Сульфат магния.xlsx",
    "2835240000": "09_Монокалийфосфат.xlsx",
    "3105100": "Экспорт NPK ВРУ.xlsx",
    "3105200": "Экспорт NPK ВРУ.xlsx",
    "3105201": "Экспорт NPK ВРУ.xlsx",
    "3105209": "Экспорт NPK ВРУ.xlsx",
    "3105908": "Экспорт NPK ВРУ.xlsx",
    "3105400001": "Экспорт МАФ 12-61.xlsx",
}

# Соответствие файлов и листов
sheet_mapping = {
    "01_Калиевая селитра.xlsx": ["Данные"],
    "02_Формиаты.xlsx": ["Данные"],
    "03_Нитрат кальция.xlsx": ["Данные"],
    "05_Нитрит натрия.xlsx": ["Данные"],
    "06_МАФ.xlsx": ["Данные"],
    "07_NPK(S) ВРУ.xlsx": ["Данные для объемов", "Данные для цен"],
    "08_Сульфат магния.xlsx": ["Данные"],
    "09_Монокалийфосфат.xlsx": ["Данные"],
    "Экспорт NPK ВРУ.xlsx": ["Данные"],
    "Экспорт МАФ 12-61.xlsx": ["Данные"],
    "Экспорт Монокалийфосфата.xlsx": ["Лист1"],
    "Экспорт Сульфата магния.xlsx": ["Данные"]
}


def get_sheet_columns(filename, sheet_name):
    #Получает заголовки колонок из указанного листа Excel
    try:
        df = pd.read_excel(filename, sheet_name=sheet_name, nrows=0)
        return df.columns.tolist()
    except Exception as e:
        print(f"❌ Ошибка чтения заголовков из '{filename}', лист '{sheet_name}': {e}")
        return []


def append_df_to_excel(filename, df, sheet_name='Sheet1'):
    #Добавляет DataFrame в конец указанного листа Excel
    if os.path.exists(filename):
        book = load_workbook(filename)
        if sheet_name in book.sheetnames:
            ws = book[sheet_name]
        else:
            ws = book.create_sheet(sheet_name)
    else:
        from openpyxl import Workbook
        book = Workbook()
        book.remove(book.active)
        ws = book.create_sheet(sheet_name)

    old_max_row = ws.max_row

    print(f"\n📄 Файл: {filename}, Лист: {sheet_name}")
    print(f"📏 Старый размер: {old_max_row} строк")
    print(f"🆕 Новый размер: {len(df)} строк")
    print(f"🧮 Общий размер после добавления: {old_max_row + len(df)} строк")

    startrow = old_max_row + 1

    # Добавляем заголовки, если файл пустой
    if old_max_row == 0 and not df.empty:
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=startrow, column=col_idx, value=col_name)
        startrow += 1

    # Записываем данные
    for row in dataframe_to_rows(df, index=False, header=False):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=startrow, column=c_idx, value=value)
        startrow += 1

    book.save(filename)


def process_data():
    # 1. Проверка и подготовка
    if not os.path.exists(input_file):
        print(f"❌ Ошибка: исходный файл не найден: {input_file}")
        return

    os.makedirs(output_folder, exist_ok=True)
    print(f"🔹 Папка для результатов: {output_folder}")

    # 2. Загрузка исходных данных
    source_df = pd.read_excel(input_file)
    print(f"✅ Загружен исходный файл. Записей: {len(source_df)}")

    if 'G33 (код товара по ТН ВЭД РФ)' not in source_df.columns:
        print("❌ В файле отсутствует колонка 'G33 (код товара по ТН ВЭД РФ)'")
        return

    # 3. Группировка данных
    results = {}  # <-- Объявляем результат внутри функции

    for filename, sheet_names in sheet_mapping.items():
        target_path = os.path.join(target_folder, filename)
        output_path = os.path.join(output_folder, filename)

        # Копируем файл из Target в Output, если его там нет
        if not os.path.exists(output_path) and os.path.exists(target_path):
            shutil.copy(target_path, output_path)

        # Читаем заголовки из целевого файла
        sheet_columns = {}
        for sheet in sheet_names:
            cols = get_sheet_columns(output_path, sheet)
            sheet_columns[sheet] = cols

        # Фильтруем данные по code_mapping
        matched_codes = [k for k, v in code_mapping.items() if v == filename]
        file_data = source_df[source_df['G33 (код товара по ТН ВЭД РФ)'].astype(str).isin(matched_codes)]

        if not file_data.empty:
            results[output_path] = {'sheet_columns': sheet_columns, 'data': file_data}

    # 4. Сохранение данных в нужные файлы и листы
    for output_path, data_info in results.items():
        sheet_columns = data_info['sheet_columns']
        full_data = data_info['data']

        matched_codes = [k for k, v in code_mapping.items() if v == os.path.basename(output_path)]

        print(f"\n📊 Статистика переноса данных для {os.path.basename(output_path)}:")
        total_rows = 0

        for sheet_name, columns in sheet_columns.items():
            # Формируем DataFrame с тем же порядком колонок, как в шаблоне
            filtered_data = pd.DataFrame()

            for col in columns:
                if col in source_df.columns:
                    filtered_data[col] = full_data[col]
                else:
                    filtered_data[col] = None  # Оставляем пустую колонку

            if not filtered_data.empty:
                append_df_to_excel(output_path, filtered_data, sheet_name=sheet_name)
                total_rows += len(filtered_data)

                print(f"📌 Лист '{sheet_name}'")
                print(f"   • Найдено записей: {len(filtered_data)}")
                missing_cols = [col for col in columns if col not in source_df.columns]
                if missing_cols:
                    print(f"   ⚠️ Отсутствующие колонки в input: {missing_cols}")

        print(f"📦 Всего добавлено строк: {total_rows}")
        matched_codes_str = "', '".join(matched_codes)
        print(f"🏷️  По кодам ТН ВЭД: '{matched_codes_str}'")
        print(f"💾 Сохранен файл: {os.path.basename(output_path)}")

    print("\n✅ Обработка завершена!")


if __name__ == "__main__":
    process_data()
